from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, BackgroundTasks
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, validator
from typing import List, Optional, Dict, Any
import os
import logging
import uuid
import io
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
from decimal import Decimal
import re
from pathlib import Path
from dotenv import load_dotenv
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from io import BytesIO
from enum import Enum
import bcrypt
import jwt
from fastapi.middleware.cors import CORSMiddleware

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# FastAPI app
app = FastAPI(title="Activus Invoice Management System", version="1.0.0")
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer()
SECRET_KEY = "activus_secret_key_2024"

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Models
class UserRole(str, Enum):
    SUPER_ADMIN = "super_admin"
    INVOICE_CREATOR = "invoice_creator"
    REVIEWER = "reviewer"
    APPROVER = "approver"
    CLIENT = "client"

class InvoiceType(str, Enum):
    PROFORMA = "proforma"
    TAX_INVOICE = "tax_invoice"
    RUNNING_ACCOUNT = "running_account"

class InvoiceStatus(str, Enum):
    DRAFT = "draft"
    PENDING_REVIEW = "pending_review"
    REVIEWED = "reviewed"
    APPROVED = "approved"
    PAID = "paid"

class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    password_hash: str
    role: UserRole
    company_name: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=datetime.utcnow)

class UserCreate(BaseModel):
    email: str
    password: str
    role: UserRole
    company_name: Optional[str] = None

class UserLogin(BaseModel):
    email: str
    password: str

class ProjectMetadata(BaseModel):
    project_name: Optional[str] = None
    architect: Optional[str] = None
    client: Optional[str] = None
    location: Optional[str] = None
    date: Optional[str] = None

class BOQItem(BaseModel):
    serial_number: str
    description: str
    unit: str
    quantity: float
    rate: float
    amount: float
    category: Optional[str] = None

class ClientInfo(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    gst_no: Optional[str] = None
    bill_to_address: str
    ship_to_address: Optional[str] = None
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class Project(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_name: str
    architect: str
    client_id: str
    client_name: str
    metadata: ProjectMetadata
    boq_items: List[BOQItem]
    total_project_value: float
    advance_received: float = 0.0
    pending_payment: float = 0.0
    created_by: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class Invoice(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: str
    project_id: str
    project_name: str
    client_id: str
    client_name: str
    invoice_type: InvoiceType
    ra_sequence: Optional[int] = None  # For running account invoices (RA1, RA2, etc.)
    include_gst: bool = True  # Whether to include GST in this invoice
    items: List[BOQItem]
    subtotal: float
    gst_amount: float
    total_amount: float
    status: InvoiceStatus = InvoiceStatus.DRAFT
    created_by: str
    reviewed_by: Optional[str] = None
    approved_by: Optional[str] = None
    invoice_date: datetime = Field(default_factory=datetime.utcnow)
    due_date: Optional[datetime] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class ActivityLog(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_email: str
    user_role: str
    action: str
    description: str
    project_id: Optional[str] = None
    invoice_id: Optional[str] = None
    timestamp: datetime = Field(default_factory=datetime.utcnow)

class CompanySettings(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_name: str = "Activus"
    logo_path: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    website: Optional[str] = None
    updated_at: datetime = Field(default_factory=datetime.utcnow)

# Excel Parser Class
class ExcelParser:
    def __init__(self):
        self.metadata_patterns = {
            'project_name': [r'project\s*name', r'project\s*:', r'job\s*name'],
            'architect': [r'architect', r'architect\s*name', r'architect\s*:'],
            'client': [r'client', r'client\s*name', r'client\s*:'],
            'location': [r'location', r'site', r'address'],
            'date': [r'date', r'project\s*date']
        }
    
    async def parse_excel_file(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            worksheet = self._select_worksheet(workbook)
            
            if not worksheet:
                raise ValueError("No valid worksheet found in the Excel file")
            
            # Extract metadata
            metadata = self._extract_metadata(worksheet)
            
            # Extract BOQ items
            items = self._extract_boq_items(worksheet)
            
            if not items:
                logger.warning(f"No BOQ items found in file {filename}")
                # Still allow processing with empty items
            
            # Calculate totals with validation
            total_value = 0.0
            try:
                total_value = sum(item['amount'] for item in items if item.get('amount'))
            except (TypeError, ValueError) as e:
                logger.warning(f"Error calculating total value: {e}")
                total_value = 0.0
            
            return {
                'metadata': metadata,
                'items': items,
                'total_value': total_value,
                'filename': filename,
                'items_count': len(items)
            }
        except Exception as e:
            logger.error(f"Error parsing Excel file {filename}: {e}")
            raise HTTPException(
                status_code=400, 
                detail=f"Error parsing Excel file: {str(e)}. Please ensure the file contains BOQ data in the expected format."
            )
    
    def _select_worksheet(self, workbook: openpyxl.Workbook):
        """Select the appropriate worksheet containing BOQ data"""
        # Priority order for worksheet selection
        preferred_names = ['boq', 'bill of quantities', 'quantities', 'estimate', 'summary', 'sheet1', 'main']
        
        # First try to find by preferred names
        for name in preferred_names:
            for sheet_name in workbook.sheetnames:
                if name.lower() in sheet_name.lower():
                    return workbook[sheet_name]
        
        # If no preferred name found, return the first non-empty sheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            if worksheet.max_row > 1 and worksheet.max_column > 1:
                return worksheet
        
        # Return active worksheet as last resort
        return workbook.active
    
    def _extract_metadata(self, worksheet) -> Dict[str, Any]:
        metadata = {}
        
        # Search first 20 rows for metadata
        for row_idx in range(1, min(21, worksheet.max_row + 1)):
            for col_idx in range(1, min(10, worksheet.max_column + 1)):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    self._extract_metadata_field(cell.value, worksheet, row_idx, col_idx, metadata)
        
        return metadata
    
    def _extract_metadata_field(self, cell_value: str, worksheet, row_idx: int, col_idx: int, metadata: Dict):
        cell_lower = cell_value.lower().strip()
        
        for field, patterns in self.metadata_patterns.items():
            for pattern in patterns:
                if re.search(pattern, cell_lower):
                    value = self._find_adjacent_value(worksheet, row_idx, col_idx)
                    if value:
                        metadata[field] = value
                        break
    
    def _find_adjacent_value(self, worksheet, row_idx: int, col_idx: int) -> Optional[str]:
        # Check same cell after colon/dash
        current_cell = worksheet.cell(row=row_idx, column=col_idx)
        if current_cell.value and (':' in str(current_cell.value) or '-' in str(current_cell.value)):
            parts = re.split(r'[:\-]', str(current_cell.value), 1)
            if len(parts) > 1 and parts[1].strip():
                return parts[1].strip()
        
        # Check right cell
        right_cell = worksheet.cell(row=row_idx, column=col_idx + 1)
        if right_cell.value and str(right_cell.value).strip():
            return str(right_cell.value).strip()
        
        return None
    
    def _extract_boq_items(self, worksheet) -> List[Dict]:
        items = []
        header_row = self._find_header_row(worksheet)
        
        if not header_row:
            return items
        
        column_mapping = self._map_columns(worksheet, header_row)
        
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            row_data = self._extract_row_data(worksheet, row_idx, column_mapping)
            
            if self._is_valid_item_row(row_data):
                quantity = self._safe_float_conversion(row_data.get('quantity')) or 0.0
                rate = self._safe_float_conversion(row_data.get('rate')) or 0.0
                amount = self._safe_float_conversion(row_data.get('amount'))
                
                # Calculate amount if not provided
                if amount is None or amount == 0:
                    amount = quantity * rate
                
                items.append({
                    'serial_number': str(row_data.get('serial', row_idx - header_row)),
                    'description': str(row_data.get('description', '')).strip(),
                    'unit': str(row_data.get('unit', 'nos')).strip() or 'nos',
                    'quantity': quantity,
                    'rate': rate,
                    'amount': amount
                })
        
        return items
    
    def _find_header_row(self, worksheet) -> Optional[int]:
        header_keywords = ['description', 'quantity', 'rate', 'amount', 'item', 'particular']
        
        for row_idx in range(1, min(30, worksheet.max_row + 1)):
            row_cells = [str(worksheet.cell(row=row_idx, column=col).value or '').lower() for col in range(1, min(20, worksheet.max_column + 1))]
            row_text = ' '.join(row_cells)
            
            matches = sum(1 for keyword in header_keywords if keyword in row_text)
            if matches >= 3:
                return row_idx
        
        return None
    
    def _map_columns(self, worksheet, header_row: int) -> Dict[str, int]:
        column_mapping = {}
        
        for col_idx in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=header_row, column=col_idx).value
            if cell_value:
                cell_lower = str(cell_value).lower().strip()
                
                if any(h in cell_lower for h in ['s.no', 'sr.no', 'serial', 'sl']):
                    column_mapping['serial'] = col_idx
                elif any(h in cell_lower for h in ['description', 'item', 'particular']):
                    column_mapping['description'] = col_idx
                elif any(h in cell_lower for h in ['unit', 'uom']):
                    column_mapping['unit'] = col_idx
                elif any(h in cell_lower for h in ['quantity', 'qty']):
                    column_mapping['quantity'] = col_idx
                elif any(h in cell_lower for h in ['rate', 'price', 'unit rate']):
                    column_mapping['rate'] = col_idx
                elif any(h in cell_lower for h in ['amount', 'total']):
                    column_mapping['amount'] = col_idx
        
        return column_mapping
    
    def _extract_row_data(self, worksheet, row_idx: int, column_mapping: Dict[str, int]) -> Dict:
        row_data = {}
        
        for field, col_idx in column_mapping.items():
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            row_data[field] = cell_value
        
        return row_data
    
    def _is_valid_item_row(self, row_data: Dict) -> bool:
        description = row_data.get('description', '')
        if not description or not isinstance(description, str) or len(description.strip()) < 3:
            return False
        
        # Must have at least description and one numeric field
        numeric_fields = ['quantity', 'rate', 'amount']
        has_numeric = any(
            self._safe_float_conversion(row_data.get(field)) is not None
            for field in numeric_fields
        )
        
        return has_numeric
    
    def _safe_float_conversion(self, value) -> Optional[float]:
        """Safely convert value to float, handling None and invalid values"""
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return float(value)
        
        if isinstance(value, str):
            # Remove common formatting characters
            cleaned = value.strip().replace(',', '').replace('₹', '').replace('Rs.', '').replace('Rs', '')
            if not cleaned:
                return None
            try:
                return float(cleaned)
            except ValueError:
                return None
        
        return None

# PDF Generator Class
class PDFGenerator:
    def __init__(self):
        self.page_size = A4
        self.margin = 20 * mm
        
    async def generate_invoice_pdf(self, invoice: Invoice, project: Project, client: ClientInfo, company_settings: Optional[dict] = None):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            spaceAfter=30,
            textColor=colors.HexColor('#1f4e79')
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#1f4e79')
        )
        
        # Get company settings for logo
        if company_settings:
            logo_path = company_settings.get('logo_path')
        else:
            # Fallback to default if company settings are not available
            logo_path = None
        
        # Header with logo (if available) and company info
        header_data = []
        
        # Try to add logo if available
        logo_cell_content = []
        if logo_path and os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=80, height=60)
                logo_cell_content.append(logo)
            except:
                # If logo fails to load, add company name instead
                logo_cell_content.append(Paragraph("ACTIVUS", title_style))
        else:
            logo_cell_content.append(Paragraph("ACTIVUS", title_style))
        
        company_info = [
            Paragraph("ACTIVUS CORPORATION", styles['Normal']),
            Paragraph("Invoice Management System", styles['Normal']),
            Paragraph("Email: info@activus.com", styles['Normal'])
        ]
        
        header_data.append([logo_cell_content, company_info])
        
        header_table = Table(header_data, colWidths=[100*mm, 100*mm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 20))
        
        # Invoice Title
        invoice_title = f"{invoice.invoice_type.upper()} INVOICE"
        if invoice.invoice_type == InvoiceType.RUNNING_ACCOUNT:
            invoice_title = f"RUNNING ACCOUNT INVOICE - RA{invoice.ra_sequence}"
        
        elements.append(Paragraph(invoice_title, title_style))
        elements.append(Spacer(1, 20))
        
        # Invoice Details
        details_data = [
            ['Invoice Number:', invoice.invoice_number, 'Date:', invoice.invoice_date.strftime('%d/%m/%Y')],
            ['Project:', project.project_name, 'Client:', client.name],
            ['Architect:', project.architect, '', '']
        ]
        
        details_table = Table(details_data, colWidths=[40*mm, 60*mm, 30*mm, 50*mm])
        details_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(details_table)
        elements.append(Spacer(1, 20))
        
        # Client Information
        elements.append(Paragraph("Bill To:", header_style))
        elements.append(Paragraph(client.name, styles['Normal']))
        elements.append(Paragraph(client.bill_to_address, styles['Normal']))
        if client.gst_no:
            elements.append(Paragraph(f"GST No: {client.gst_no}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Items Table
        items_data = [['S.No', 'Description', 'Unit', 'Qty', 'Rate', 'Amount']]
        
        for item in invoice.items:
            items_data.append([
                item.serial_number,
                item.description,
                item.unit,
                str(item.quantity),
                f"₹{item.rate:,.2f}",
                f"₹{item.amount:,.2f}"
            ])
        
        items_table = Table(items_data, colWidths=[15*mm, 80*mm, 20*mm, 20*mm, 30*mm, 35*mm])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4e79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(items_table)
        elements.append(Spacer(1, 20))
        
        # Totals
        totals_data = [
            ['Subtotal:', f"₹{invoice.subtotal:,.2f}"]
        ]
        
        if invoice.include_gst:
            totals_data.append(['GST (18%):', f"₹{invoice.gst_amount:,.2f}"])
        
        totals_data.append(['Total Amount:', f"₹{invoice.total_amount:,.2f}"])
        
        totals_table = Table(totals_data, colWidths=[50*mm, 50*mm])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(totals_table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

# Authentication functions
async def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')

async def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

async def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        'user_id': user_id,
        'email': email,
        'role': role,
        'exp': datetime.utcnow() + timedelta(days=7)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm='HS256')

async def verify_token(token: str) -> Dict:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    payload = await verify_token(credentials.credentials)
    user = await db.users.find_one({"id": payload["user_id"]})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

async def log_activity(user_id: str, user_email: str, user_role: str, action: str, description: str, 
                      project_id: Optional[str] = None, invoice_id: Optional[str] = None):
    log_entry = ActivityLog(
        user_id=user_id,
        user_email=user_email,
        user_role=user_role,
        action=action,
        description=description,
        project_id=project_id,
        invoice_id=invoice_id
    )
    await db.activity_logs.insert_one(log_entry.dict())

# Initialize super admin
async def init_super_admin():
    super_admin_email = "brightboxm@gmail.com"
    existing_user = await db.users.find_one({"email": super_admin_email})
    
    if not existing_user:
        password_hash = await hash_password("admin123")
        super_admin = User(
            email=super_admin_email,
            password_hash=password_hash,
            role=UserRole.SUPER_ADMIN,
            company_name="Activus Industrial Design & Build"
        )
        await db.users.insert_one(super_admin.dict())
        logger.info("Super admin created successfully")

# API Routes
@api_router.post("/auth/login")
async def login(user_data: UserLogin):
    try:
        user = await db.users.find_one({"email": user_data.email, "is_active": True})
        if not user:
            raise HTTPException(status_code=401, detail="Invalid email or password")
        
        if not await verify_password(user_data.password, user["password_hash"]):
            raise HTTPException(status_code=401, detail="Invalid email or password")
        
        token = await create_token(user["id"], user["email"], user["role"])
        
        await log_activity(
            user["id"], user["email"], user["role"], 
            "login", f"User logged in successfully"
        )
        
        return {
            "access_token": token,
            "token_type": "bearer",
            "user": {
                "id": user["id"],
                "email": user["email"],
                "role": user["role"],
                "company_name": user["company_name"]
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Login error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

@api_router.post("/auth/register")
async def register(user_data: UserCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only super admin can create users")
    
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    password_hash = await hash_password(user_data.password)
    new_user = User(
        email=user_data.email,
        password_hash=password_hash,
        role=user_data.role,
        company_name=user_data.company_name
    )
    
    await db.users.insert_one(new_user.dict())
    
    await log_activity(
        current_user["id"], current_user["email"], current_user["role"],
        "user_created", f"Created new user: {user_data.email} with role: {user_data.role}"
    )
    
    return {"message": "User created successfully", "user_id": new_user.id}

@api_router.post("/upload-boq")
async def upload_boq(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    # Validate file type
    allowed_content_types = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel',
        'application/vnd.ms-excel.sheet.macroEnabled.12'
    ]
    
    allowed_extensions = ['.xlsx', '.xlsm', '.xls']
    
    if file.content_type not in allowed_content_types:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid file type. Please upload an Excel file (.xlsx, .xls, .xlsm). Received: {file.content_type}"
        )
    
    if not any(file.filename.lower().endswith(ext) for ext in allowed_extensions):
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid file extension. Please upload an Excel file with .xlsx, .xls, or .xlsm extension"
        )
    
    # Check file size (max 10MB)
    file_content = await file.read()
    if len(file_content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large. Maximum size is 10MB")
    
    if len(file_content) == 0:
        raise HTTPException(status_code=400, detail="Empty file uploaded")
    
    try:
        parser = ExcelParser()
        parsed_data = await parser.parse_excel_file(file_content, file.filename)
        
        await log_activity(
            current_user["id"], current_user["email"], current_user["role"],
            "boq_uploaded", f"Successfully uploaded and parsed BOQ file: {file.filename}"
        )
        
        return parsed_data
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"BOQ parsing error for file {file.filename}: {str(e)}")
        raise HTTPException(
            status_code=422, 
            detail=f"Failed to parse Excel file. Please ensure it's a valid BOQ format. Error: {str(e)}"
        )

@api_router.post("/upload-logo")
async def upload_logo(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload company logo"""
    if current_user["role"] not in ["super_admin"]:
        raise HTTPException(status_code=403, detail="Only super admin can upload logo")
    
    # Validate file type
    allowed_types = ["image/jpeg", "image/jpg", "image/png", "image/gif"]
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type. Please upload an image file. Received: {file.content_type}"
        )
    
    # Validate file extension
    allowed_extensions = ['.jpg', '.jpeg', '.png', '.gif']
    file_extension = os.path.splitext(file.filename)[1].lower()
    if file_extension not in allowed_extensions:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file extension. Please upload an image file with .jpg, .jpeg, .png, or .gif extension"
        )
    
    try:
        # Read file content
        file_content = await file.read()
        if len(file_content) == 0:
            raise HTTPException(status_code=400, detail="Empty file uploaded")
        
        # Create uploads directory if it doesn't exist
        upload_dir = "uploads/logos"
        os.makedirs(upload_dir, exist_ok=True)
        
        # Generate unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"logo_{timestamp}{file_extension}"
        file_path = os.path.join(upload_dir, filename)
        
        # Save file
        with open(file_path, "wb") as f:
            f.write(file_content)
        
        # Update or create company settings
        company_settings = await db.company_settings.find_one({})
        if company_settings:
            # Update existing settings
            await db.company_settings.update_one(
                {"_id": company_settings["_id"]},
                {"$set": {"logo_path": file_path, "updated_at": datetime.utcnow()}}
            )
        else:
            # Create new settings
            new_settings = CompanySettings(logo_path=file_path)
            await db.company_settings.insert_one(new_settings.dict())
        
        await log_activity(
            current_user["id"], current_user["email"], current_user["role"],
            "logo_uploaded", f"Uploaded company logo: {filename}"
        )
        
        return {"message": "Logo uploaded successfully", "logo_path": file_path}
        
    except Exception as e:
        logger.error(f"Error uploading logo: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to upload logo")

@api_router.get("/company-settings")
async def get_company_settings(current_user: dict = Depends(get_current_user)):
    """Get company settings including logo"""
    settings = await db.company_settings.find_one({})
    if settings:
        return CompanySettings(**settings)
    else:
        # Return default settings
        return CompanySettings()

@api_router.put("/company-settings")
async def update_company_settings(
    settings: CompanySettings,
    current_user: dict = Depends(get_current_user)
):
    """Update company settings"""
    if current_user["role"] not in ["super_admin"]:
        raise HTTPException(status_code=403, detail="Only super admin can update company settings")
    
    existing_settings = await db.company_settings.find_one({})
    if existing_settings:
        # Update existing settings
        await db.company_settings.update_one(
            {"_id": existing_settings["_id"]},
            {"$set": settings.dict()}
        )
    else:
        # Create new settings
        await db.company_settings.insert_one(settings.dict())
    
    await log_activity(
        current_user["id"], current_user["email"], current_user["role"],
        "company_settings_updated", "Updated company settings"
    )
    
    return {"message": "Company settings updated successfully"}

@api_router.post("/clients", response_model=dict)
async def create_client(client_data: ClientInfo, current_user: dict = Depends(get_current_user)):
    await db.clients.insert_one(client_data.dict())
    
    await log_activity(
        current_user["id"], current_user["email"], current_user["role"],
        "client_created", f"Created client: {client_data.name}"
    )
    
    return {"message": "Client created successfully", "client_id": client_data.id}

@api_router.get("/clients", response_model=List[ClientInfo])
async def get_clients(current_user: dict = Depends(get_current_user)):
    clients = await db.clients.find().to_list(1000)
    return [ClientInfo(**client) for client in clients]

@api_router.post("/projects", response_model=dict)
async def create_project(project_data: Project, current_user: dict = Depends(get_current_user)):
    try:
        # Set additional fields
        project_data.created_by = current_user["id"]
        project_data.pending_payment = project_data.total_project_value - project_data.advance_received
        project_data.updated_at = datetime.utcnow()
        
        # Validate BOQ items are properly formed
        if not project_data.boq_items:
            raise HTTPException(status_code=400, detail="BOQ items cannot be empty")
        
        # Insert into database
        await db.projects.insert_one(project_data.dict())
        
        # Log activity
        await log_activity(
            current_user["id"], current_user["email"], current_user["role"],
            "project_created", f"Created project: {project_data.project_name}",
            project_id=project_data.id
        )
        
        return {"message": "Project created successfully", "project_id": project_data.id}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating project: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create project: {str(e)}")

@api_router.get("/projects", response_model=List[Project])
async def get_projects(current_user: dict = Depends(get_current_user)):
    projects = await db.projects.find().to_list(1000)
    return [Project(**project) for project in projects]

@api_router.get("/projects/{project_id}", response_model=Project)
async def get_project(project_id: str, current_user: dict = Depends(get_current_user)):
    try:
        if not project_id or len(project_id.strip()) == 0:
            raise HTTPException(status_code=400, detail="Project ID is required")
        
        project = await db.projects.find_one({"id": project_id})
        if not project:
            raise HTTPException(status_code=404, detail=f"Project with ID {project_id} not found")
        
        return Project(**project)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error retrieving project {project_id}: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

@api_router.post("/invoices", response_model=dict)
async def create_invoice(invoice_data: Invoice, current_user: dict = Depends(get_current_user)):
    # Generate invoice number
    invoice_count = await db.invoices.count_documents({})
    
    # Handle running account invoice numbering
    if invoice_data.invoice_type == InvoiceType.RUNNING_ACCOUNT:
        # Find the highest RA sequence for this project
        existing_ra_invoices = await db.invoices.find({
            "project_id": invoice_data.project_id,
            "invoice_type": "running_account"
        }).sort("ra_sequence", -1).limit(1).to_list(1)
        
        if existing_ra_invoices:
            invoice_data.ra_sequence = existing_ra_invoices[0]["ra_sequence"] + 1
        else:
            invoice_data.ra_sequence = 1
            
        invoice_data.invoice_number = f"RA{invoice_data.ra_sequence}-{datetime.now().year}-{invoice_count + 1:04d}"
        
        # GST logic for running account: RA1 might skip GST, RA2+ should include GST
        if invoice_data.ra_sequence == 1 and not invoice_data.include_gst:
            # RA1 with GST skipped
            invoice_data.include_gst = False
        else:
            # RA2+ should include GST
            invoice_data.include_gst = True
    else:
        invoice_data.invoice_number = f"INV-{datetime.now().year}-{invoice_count + 1:04d}"
    
    invoice_data.created_by = current_user["id"]
    
    # Calculate totals
    invoice_data.subtotal = sum(item.amount for item in invoice_data.items)
    if invoice_data.include_gst:
        invoice_data.gst_amount = invoice_data.subtotal * 0.18  # 18% GST
    else:
        invoice_data.gst_amount = 0.0
    invoice_data.total_amount = invoice_data.subtotal + invoice_data.gst_amount
    
    await db.invoices.insert_one(invoice_data.dict())
    
    await log_activity(
        current_user["id"], current_user["email"], current_user["role"],
        "invoice_created", f"Created {invoice_data.invoice_type} invoice: {invoice_data.invoice_number}",
        project_id=invoice_data.project_id, invoice_id=invoice_data.id
    )
    
    return {"message": "Invoice created successfully", "invoice_id": invoice_data.id}

@api_router.get("/invoices", response_model=List[Invoice])
async def get_invoices(current_user: dict = Depends(get_current_user)):
    invoices = await db.invoices.find().to_list(1000)
    return [Invoice(**invoice) for invoice in invoices]

@api_router.get("/invoices/{invoice_id}/pdf")
async def download_invoice_pdf(invoice_id: str, current_user: dict = Depends(get_current_user)):
    try:
        invoice = await db.invoices.find_one({"id": invoice_id})
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        project = await db.projects.find_one({"id": invoice["project_id"]})
        if not project:
            raise HTTPException(status_code=404, detail="Related project not found")
            
        client = await db.clients.find_one({"id": invoice["client_id"]})
        if not client:
            raise HTTPException(status_code=404, detail="Related client not found")
        
        # Get company settings for logo
        company_settings = await db.company_settings.find_one({})
        
        pdf_generator = PDFGenerator()
        pdf_buffer = await pdf_generator.generate_invoice_pdf(
            Invoice(**invoice),
            Project(**project),
            ClientInfo(**client),
            company_settings
        )
        
        await log_activity(
            current_user["id"], current_user["email"], current_user["role"],
            "invoice_downloaded", f"Downloaded PDF for invoice: {invoice['invoice_number']}",
            invoice_id=invoice_id
        )
        
        return StreamingResponse(
            io.BytesIO(pdf_buffer.read()),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=invoice_{invoice['invoice_number']}.pdf"}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading invoice PDF {invoice_id}: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate PDF")

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    total_projects = await db.projects.count_documents({})
    total_invoices = await db.invoices.count_documents({})
    
    # Calculate project financial stats
    project_pipeline = [
        {"$group": {
            "_id": None,
            "total_project_value": {"$sum": "$total_project_value"},
            "total_advance": {"$sum": "$advance_received"}
        }}
    ]
    
    project_stats = await db.projects.aggregate(project_pipeline).to_list(1)
    total_project_value = project_stats[0]["total_project_value"] if project_stats else 0
    total_advance = project_stats[0]["total_advance"] if project_stats else 0
    
    # Calculate invoice financial stats
    invoice_pipeline = [
        {"$group": {
            "_id": None,
            "total_invoiced_value": {"$sum": "$total_amount"}
        }}
    ]
    
    invoice_stats = await db.invoices.aggregate(invoice_pipeline).to_list(1)
    total_invoiced = invoice_stats[0]["total_invoiced_value"] if invoice_stats else 0
    
    # Calculate payment status breakdown
    payment_status_pipeline = [
        {"$group": {
            "_id": "$status",
            "count": {"$sum": 1},
            "amount": {"$sum": "$total_amount"}
        }}
    ]
    
    payment_breakdown = await db.invoices.aggregate(payment_status_pipeline).to_list(10)
    
    # Calculate pending payment (invoiced but not paid)
    pending_payment = 0
    paid_amount = 0
    draft_amount = 0
    
    for item in payment_breakdown:
        if item["_id"] in ["draft", "pending_review", "reviewed", "approved"]:
            pending_payment += item["amount"]
        elif item["_id"] == "paid":
            paid_amount += item["amount"]
        if item["_id"] == "draft":
            draft_amount += item["amount"]
    
    # Calculate project completion percentage
    completion_percentage = (total_invoiced / total_project_value * 100) if total_project_value > 0 else 0
    
    return {
        "total_projects": total_projects,
        "total_invoices": total_invoices,
        "total_project_value": total_project_value,
        "total_invoiced_value": total_invoiced,
        "total_advance_received": total_advance,
        "pending_payment": pending_payment,
        "paid_amount": paid_amount,
        "draft_amount": draft_amount,
        "completion_percentage": round(completion_percentage, 2),
        "payment_breakdown": {item["_id"]: {"count": item["count"], "amount": item["amount"]} for item in payment_breakdown}
    }

@api_router.get("/activity-logs")
async def get_activity_logs(
    skip: int = 0, 
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Only super admin can view logs")
    
    logs = await db.activity_logs.find().skip(skip).limit(limit).sort("timestamp", -1).to_list(limit)
    return [ActivityLog(**log) for log in logs]

# Include router
app.include_router(api_router)

@app.on_event("startup")
async def startup_event():
    await init_super_admin()
    logger.info("Application started successfully")

@app.on_event("shutdown")
async def shutdown_event():
    client.close()
    logger.info("Application shutdown")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)